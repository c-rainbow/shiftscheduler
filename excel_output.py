
import openpyxl  

from openpyxl import styles
from openpyxl.styles import fills

from openpyxl.formatting import rule as xlrule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import cell as xlcell
import date_util
import datetime
import data
import collections



SoftwareConfig = collections.namedtuple(
    'SoftwareConfig', ['start_date', 'end_date', 'num_person'])


# Config values for each person. The order must be preserved
CONFIG_PERSON_HEADER_NAMES = (
    '최대 근무일 (연속)',
    '최대 나이트 (연속)',
    '최소 근무일 (전체)',
    '최대 근무일 (전체)',
)


# Config values for each date. The order must be preserved
CONFIG_DATE_HEADER_NAMES = (
    '데이 근무자 수',
    '이브닝 근무자 수',
    '나이트 근무자 수',
)


# Background colors to use for timetable
COLOR_MELON = 'fdbcb4'  # Peach-like soft red
COLOR_PASTEL_YELLOW = 'fdfd96'
COLOR_AERO_BLUE = 'c9ffe5'
COLOR_LIGHT_GRAY = 'bbbbbb'


# Background fills to user for timetable
DAY_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_MELON, end_color=COLOR_MELON)
EVENING_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_AERO_BLUE, end_color=COLOR_AERO_BLUE)
NIGHT_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_PASTEL_YELLOW, end_color=COLOR_PASTEL_YELLOW)
OFF_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY)


# If all names are unique. Prevents mistaken duplicate names
def AllNamesUnique(names):
    return len(names) == len(set(names))


# ws: Excel sheet
# schedule: data.Schedule, for start_date and end_date
# assignment: data.Assignment, for assignment
def CreateOutputTimetable(ws, schedule, assignments, row_offset=0, col_offset=0):
    
    names = assignments.GetNames()
    start_date = schedule.start_date
    end_date = schedule.end_date
    
    # TODO: Create custom exceptions
    if not AllNamesUnique(names):
        raise Exception('Duplicate names')
    if start_date > end_date:
        raise Exception('Start date is later than end date')
    
    # Fill the head rows with people's names
    # Note that row/col indexes are 1-based in Excel    
    names = sorted(names)
    for i, name in enumerate(names):
        c = ws.cell(row=i+2+row_offset, column=1+col_offset)
        c.value = name
    
    # Fill the head columns with dates
    all_dates = list(date_util.GenerateAllDates(start_date, end_date))
    for i, work_date in enumerate(all_dates):
        c = ws.cell(row=1+row_offset, column=i+2+col_offset)
        # TODO: How to format the date?
        c.value = work_date
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(i+2+col_offset)].width = 12  # 10 for date

    
    start_row = 2 + row_offset
    start_col = 2 + col_offset
    end_row = start_row + len(names) - 1
    end_col = start_col + len(all_dates) - 1


    # Fill shifts from assignments
    for row_index in range(start_row, end_row+1):
        name = ws.cell(row=row_index, column=1+col_offset).value
        for col_index in range(start_col, end_col+1):
            work_date = ws.cell(row=1+row_offset, column=col_index).value
            cell = ws.cell(row=row_index, column=col_index)
            print(type(work_date))
            shift_type = assignments.GetAssignment(work_date, name)
            
            # If shift exists, write the short name.
            # Note that this does not write Off-shift to Excel
            if shift_type is not None:
                cell.value = shift_type.ShortName()         

    # Fill the rest cells with 'Off' shifts
    for row_index in range(start_row, end_row+1):
        for col_index in range(start_col, end_col+1):
            cell = ws.cell(row=row_index, column=col_index)
            if cell.value is None:
                cell.value = data.ShiftType.Off.ShortName()

    # Add conditional formatting to cells
    cell_range = '%s%d:%s%d' % (
        xlcell.get_column_letter(start_col), start_row,
        xlcell.get_column_letter(end_col), end_row
    )

    # Add background colors based on value. Values are case-insensitive
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"D"'], fill=DAY_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"E"'], fill=EVENING_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"N"'], fill=NIGHT_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"O"'], fill=OFF_FILL))



def CreateOutputPersonConfig(ws, schedule, assignments, row_offset=0, col_offset=0):
    # Fill the head rows with people's names
    # Note that row/col indexes are 1-based in Excel    
    names = assignments.GetNames()
    for i, name in enumerate(names):
        c = ws.cell(row=i+2+row_offset, column=1+col_offset)
        c.value = name

    # Fill the head columns with config names
    for i, config_display_name in enumerate(CONFIG_PERSON_HEADER_NAMES):
        c = ws.cell(row=1+row_offset, column=i+2+col_offset)
        c.value = config_display_name
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(i+2+col_offset)].width = len(config_display_name)*1.7

    # Fill config values from schedule
    start_row = 2 + row_offset
    start_col = 2 + col_offset
    end_row = start_row + len(names) - 1

    for row_index in range(start_row, end_row+1):
        name = ws.cell(row=row_index, column=col_offset+1).value
        constraint = schedule.GetConstraintByName(name)
        
        # Write config values
        cell = ws.cell(row=row_index, column=start_col)
        cell.value = constraint.max_consecutive_workdays

        cell = ws.cell(row=row_index, column=start_col+1)
        cell.value = constraint.max_consecutive_nights

        cell = ws.cell(row=row_index, column=start_col+2)
        cell.value = constraint.min_total_workdays

        cell = ws.cell(row=row_index, column=start_col+3)
        cell.value = constraint.max_total_workdays

        

def WriteDateConfig(ws, date_constraints, config, start_row=1, start_col=1):
    date_constraint_dict = {c.work_date: c for c in date_constraints}

    # Fill the head rows with dates
    # Note that row/col indexes are 1-based in Excel    
    all_dates = list(date_util.GenerateAllDates(config.start_date, config.end_date))
    for i, work_date in enumerate(all_dates):
        c = ws.cell(row=start_row+1+i, column=start_col)
        c.value = work_date

    # Fill the head columns with config names
    for i, config_display_name in enumerate(CONFIG_DATE_HEADER_NAMES):
        c = ws.cell(row=start_row, column=start_col+1+i)
        c.value = config_display_name
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(start_col+1+i)].width = len(config_display_name)*2
    
    # Fill date configs
    for row_index in range(start_row+1, start_row+1+len(all_dates)):
        work_date = ws.cell(row=row_index, column=start_col).value
        date_constraint = date_constraint_dict.get(work_date)
        
        if date_constraint is not None:
            # Write config values.
            # This loop overwrites the work date cell with the same date. Ok for code simplicity
            for col_offset, value in enumerate(date_constraint._asdict().values()):
                cell = ws.cell(row=row_index, column=start_col+col_offset)
                cell.value = value


def WriteSoftwareConfig(ws, config, start_row=1, start_col=1):
    row_offset = 0
    # Iterate fields and values in namedtuple in order
    for name, value in config._asdict().items():
        # First column is config name
        name_cell = ws.cell(row=start_row+row_offset, column=start_col)
        name_cell.value = name
        # Second column is config value
        value_cell = ws.cell(row=start_row+row_offset, column=start_col+1)
        value_cell.value = value
        row_offset += 1


# Create Output Excel file
def GenerateOutputExcelFile(schedule, assignments, filename):
    wb = openpyxl.Workbook()
    
    ws = wb.create_sheet(title='Config')
    sconfig = SoftwareConfig(
        start_date=schedule.start_date,
        end_date=schedule.end_date,
        num_person=len(assignments.GetNames())
    )
    WriteSoftwareConfig(ws, sconfig)

    ws = wb.create_sheet(title='일정표')
    CreateOutputTimetable(wb.active, schedule, assignments)

    ws = wb.create_sheet(title='간호사별 설정')
    CreateOutputPersonConfig(ws, schedule, assignments)

    ws = wb.create_sheet(title='날짜별 설정')
    WriteDateConfig(ws, schedule, sconfig)

    wb.save(filename)
