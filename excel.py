
import openpyxl  

from openpyxl import styles
from openpyxl.styles import fills, colors

from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule



COLOR_MELON = 'fdbcb4'  # Peach-like soft red
COLOR_PASTEL_YELLOW = 'fdfd96'
COLOR_AERO_BLUE = 'c9ffe5'
COLOR_LIGHT_GRAY = '333333'

wb = openpyxl.Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A3'] = datetime.date.today()

c = ws.cell(row=4, column=5)
c.value = 'hello'
fff1 = styles.PatternFill(fill_type=fills.FILL_SOLID, start_color=COLOR_MELON, end_color=COLOR_MELON)
fff2 = styles.PatternFill(fill_type=fills.FILL_SOLID, start_color=COLOR_AERO_BLUE, end_color=COLOR_AERO_BLUE)
fff3 = styles.PatternFill(fill_type=fills.FILL_SOLID, start_color=COLOR_PASTEL_YELLOW, end_color=COLOR_PASTEL_YELLOW)
c.fill = fff3  # styles.PatternFill(patternType=fills.FILL_SOLID, start_color="fdbcb4")
#print(dir(c))#c.




ws.conditional_formatting.add('D3:E7', CellIsRule(operator='==', formula=['"D"'], fill=fff1))
ws.conditional_formatting.add('D3:E7', CellIsRule(operator='==', formula=['"E"'], fill=fff2))
ws.conditional_formatting.add('D3:E7', CellIsRule(operator='==', formula=['"N"'], fill=fff3))





# Save the file
wb.save("sample.xlsx")