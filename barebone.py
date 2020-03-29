

"""



TODO: This can be merged with excel_output.py


"""



import pandas
import openpyxl  

from openpyxl import styles
from openpyxl.styles import fills

from openpyxl.formatting import rule as xlrule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import cell as xlcell
import date_util
import datetime

# Config values for each person. The order must be preserved
CONFIG_PERSON_HEADER_NAMES = (
    '최대 근무일 (연속)',
    '최대 나이트 (연속)',
    '최소 근무일 (전체)',
    '최대 근무일 (전체)',
)


# Config values for each date. The order must be preserved
CONFIG_DATE_HEADER_NAMES = (
    '데이 근무자 수',
    '이브닝 근무자 수',
    '나이트 근무자 수',
)


# Background colors to use for timetable
COLOR_MELON = 'fdbcb4'  # Peach-like soft red
COLOR_PASTEL_YELLOW = 'fdfd96'
COLOR_AERO_BLUE = 'c9ffe5'
COLOR_LIGHT_GRAY = 'bbbbbb'


# Background fills to user for timetable
DAY_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_MELON, end_color=COLOR_MELON)
EVENING_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_AERO_BLUE, end_color=COLOR_AERO_BLUE)
NIGHT_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_PASTEL_YELLOW, end_color=COLOR_PASTEL_YELLOW)
OFF_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY)


# If all names are unique. Prevents mistaken duplicate names
def AllNamesUnique(names):
    return len(names) == len(set(names))


# ws: Excel sheet
# names: names of people. Must be all unique
# start_date: datetime.date. First date to write to Excel file
# end_date: datetime.date. Last date to write to Excel file
def CreateBareboneTimetable(ws, names, start_date, end_date, row_offset=0, col_offset=0):
    
    # TODO: Create custom exceptions
    if not AllNamesUnique(names):
        raise Exception('Duplicate names')
    if start_date > end_date:
        raise Exception('Start date is later than end date')
    
    # Fill the head rows with people's names
    # Note that row/col indexes are 1-based in Excel    
    names = sorted(names)
    for i, name in enumerate(names):
        c = ws.cell(row=i+2+row_offset, column=1+col_offset)
        c.value = name
    
    # Fill the head columns with dates
    all_dates = list(date_util.GenerateAllDates(start_date, end_date))
    for i, work_date in enumerate(all_dates):
        c = ws.cell(row=1+row_offset, column=i+2+col_offset)
        # TODO: How to format the date?
        c.value = work_date
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(i+2+col_offset)].width = 12  # 10 for date

    # Add conditional formatting to cells
    start_row = 2 + row_offset
    start_col = 2 + col_offset
    end_row = start_row + len(names) - 1
    end_col = start_col + len(all_dates) - 1
    cell_range = '%s%d:%s%d' % (
        xlcell.get_column_letter(start_col), start_row,
        xlcell.get_column_letter(end_col), end_row
    )

    # Add background colors based on value. Values are case-insensitive
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"D"'], fill=DAY_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"E"'], fill=EVENING_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"N"'], fill=NIGHT_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"O"'], fill=OFF_FILL))



def CreateBarebonePersonConfig(ws, names, row_offset=0, col_offset=0):
    # Fill the head rows with people's names
    # Note that row/col indexes are 1-based in Excel    
    names = sorted(names)
    for i, name in enumerate(names):
        c = ws.cell(row=i+2+row_offset, column=1+col_offset)
        c.value = name

    # Fill the head columns with config names
    for i, config_display_name in enumerate(CONFIG_PERSON_HEADER_NAMES):
        c = ws.cell(row=1+row_offset, column=i+2+col_offset)
        c.value = config_display_name
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(i+2+col_offset)].width = len(config_display_name)*1.7
        

def CreateBareboneDateConfig(ws, start_date, end_date, row_offset=0, col_offset=0):
    # Fill the head rows with dates
    # Note that row/col indexes are 1-based in Excel    
    all_dates = list(date_util.GenerateAllDates(start_date, end_date))
    for i, work_date in enumerate(all_dates):
        c = ws.cell(row=i+2+row_offset, column=1+col_offset)
        # TODO: How to format the date?
        c.value = work_date

    # Fill the head columns with config names
    for i, config_display_name in enumerate(CONFIG_DATE_HEADER_NAMES):
        c = ws.cell(row=1+row_offset, column=i+2+col_offset)
        c.value = config_display_name
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(i+2+col_offset)].width = len(config_display_name)*2
    

# Create Barebone Excel file
def GenerateBareboneExcelFile(names, start_date, end_date, filename):
    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = '일정표'
    CreateBareboneTimetable(wb.active, names, start_date, end_date, row_offset=0, col_offset=0)

    ws = wb.create_sheet(title='간호사별 설정')
    CreateBarebonePersonConfig(ws, names, row_offset=0, col_offset=0)

    ws = wb.create_sheet(title='날짜별 설정')
    CreateBareboneDateConfig(ws, start_date, end_date, row_offset=0, col_offset=0)

    wb.save(filename)

if __name__ == '__main__':
    GenerateBareboneExcelFile(
        ['간호사1', '간호사4', '간호사2', '간호사3'],
        datetime.date.fromisoformat('2020-02-04'),
        datetime.date.fromisoformat('2020-03-16'),
        'tmp/sample.xlsx'
    )

