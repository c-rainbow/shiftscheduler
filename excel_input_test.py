
import data
import datetime
import openpyxl  
import unittest
import excel_input


class ExcelOutputTest(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.wb = openpyxl.Workbook()

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        cls.wb = None  # Is this line necessary?

    def setUp(self):
        self.ws = ExcelOutputTest.wb.create_sheet()
        self.test_start_date = datetime.date(2020, 1, 31)
        self.test_end_date = datetime.date(2020, 2, 1)
        self.test_config = data.SoftwareConfig(
            start_date=self.test_start_date, end_date=self.test_end_date, num_person=2)

    def tearDown(self):
        ExcelOutputTest.wb.remove(self.ws)
        self.ws = None  # Is this line necessary?

    def testReadTimetableEmpty(self):
        ws = self.ws
        ws['B1'] = self.test_start_date
        ws['C1'] = self.test_end_date
        ws['A2'] = '간호사1'
        ws['A3'] = '간호사2'

        assignment_dict = excel_input.ReadTimetable(ws, self.test_config)
        self.assertEqual(len(assignment_dict), 4)
        
        self.assertIsNone(assignment_dict.get((self.test_start_date, '간호사1')))
        self.assertIsNone(assignment_dict.get((self.test_start_date, '간호사2')))
        self.assertIsNone(assignment_dict.get((self.test_end_date, '간호사1')))
        self.assertIsNone(assignment_dict.get((self.test_end_date, '간호사2')))


    def testReadTimetablePartialEmpty(self):
        ws = self.ws
        ws['B1'] = self.test_start_date
        ws['C1'] = self.test_end_date
        ws['A2'] = '간호사1'
        ws['A3'] = '간호사2'
        ws['B2'] = 'E'
        ws['B3'] = 'O'
        # C2 is intentionally empty
        ws['C3'] = 'D'

        assignment_dict = excel_input.ReadTimetable(ws, self.test_config)
        self.assertEqual(len(assignment_dict), 4)

        self.assertEqual(assignment_dict.get((self.test_start_date, '간호사1')), data.ShiftType.EVENING)
        self.assertIsNone(assignment_dict.get((self.test_end_date, '간호사1')))  # None for empty cell

        self.assertEqual(assignment_dict.get((self.test_start_date, '간호사2')), data.ShiftType.OFF)
        self.assertEqual(assignment_dict.get((self.test_end_date, '간호사2')), data.ShiftType.DAY)

    def testReadTimetable(self):
        ws = self.ws
        ws['B1'] = self.test_start_date
        ws['C1'] = self.test_end_date
        ws['A2'] = '간호사1'
        ws['A3'] = '간호사2'
        ws['A4'] = '간호사3'
        ws['B2'] = 'N'
        ws['B3'] = 'O'
        ws['C2'] = 'I'  # Invalid
        ws['C3'] = 'invalid'  # Invalid

        assignment_dict = excel_input.ReadTimetable(ws, self.test_config)
        self.assertEqual(len(assignment_dict), 4)

        self.assertEqual(assignment_dict.get((self.test_start_date, '간호사1')), data.ShiftType.NIGHT)
        self.assertEqual(assignment_dict.get((self.test_end_date, '간호사1')), data.ShiftType.UNKNOWN)

        self.assertEqual(assignment_dict.get((self.test_start_date, '간호사2')), data.ShiftType.OFF)
        self.assertEqual(assignment_dict.get((self.test_end_date, '간호사2')), data.ShiftType.UNKNOWN)
        
    # Test ReadPersonConfig()
    def testReadPersonConfig(self):
        ws = self.ws
        ws['A2'] = '간호사1'
        ws['B2'] = 5
        ws['C2'] = 3
        ws['D2'] = 24
        # E2 is intentionally empty
        ws['A3'] = '간호사2'
        ws['B3'] = 6
        ws['C3'] = 4
        ws['D3'] = 25
        ws['E3'] = 28

        person_configs = excel_input.ReadPersonConfig(ws, self.test_config)
        self.assertEqual(len(person_configs), 2)

        pc1 = person_configs[0]
        self.assertEqual(pc1.name, '간호사1')
        self.assertEqual(pc1.max_consecutive_workdays, 5)
        self.assertEqual(pc1.max_consecutive_nights, 3)
        self.assertEqual(pc1.min_total_workdays, 24)
        self.assertIsNone(pc1.max_total_workdays)

        pc2 = person_configs[1]
        self.assertEqual(pc2.name, '간호사2')
        self.assertEqual(pc2.max_consecutive_workdays, 6)
        self.assertEqual(pc2.max_consecutive_nights, 4)
        self.assertEqual(pc2.min_total_workdays, 25)
        self.assertEqual(pc2.max_total_workdays, 28)
  
    # Test ReadDateConfig()
    def testReadDateConfig(self):
        ws = self.ws
        ws['A2'] = self.test_start_date
        ws['A3'] = self.test_end_date
        ws['B2'] = 1
        ws['C2'] = 2
        ws['D2'] = 3
        # B3 is intentionally empty
        ws['C3'] = 5
        ws['D3'] = 6

        date_configs = excel_input.ReadDateConfig(ws, self.test_config)
        self.assertEqual(len(date_configs), 2)

        dc1 = date_configs[0]
        self.assertEqual(dc1.work_date, self.test_start_date) 
        self.assertEqual(dc1.num_workers_day, 1)
        self.assertEqual(dc1.num_workers_evening, 2)
        self.assertEqual(dc1.num_workers_night, 3)

        dc2 = date_configs[1]
        self.assertEqual(dc2.work_date, self.test_end_date) 
        self.assertIsNone(dc2.num_workers_day)
        self.assertEqual(dc2.num_workers_evening, 5)
        self.assertEqual(dc2.num_workers_night, 6)

    # Test ReadSoftwareConfig()
    def testReadSoftwareConfig(self):
        ws = self.ws

        # The order is intentionally different from the field order of SoftwareConfig
        ws['A1'] = 'num_person'
        ws['B1'] = 5
        ws['A2'] = 'start_date'
        ws['B2'] =  self.test_start_date
        ws['A3'] = 'end_date'
        ws['B3'] = self.test_end_date

        config = excel_input.ReadSoftwareConfig(ws)

        self.assertEqual(config.start_date, self.test_start_date)
        self.assertEqual(config.end_date, self.test_end_date)
        self.assertEqual(config.num_person, 5)

    # Test ReadSoftwareConfig() when some field is missing
    def testReadSoftwareConfigMissingField(self):
        ws = self.ws
        ws['A1'] = 'num_person'
        ws['B1'] = 5
        ws['A2'] = 'start_date'
        ws['B2'] =  self.test_start_date

        self.assertRaises(TypeError, excel_input.ReadSoftwareConfig, ws)

    # Test ReadSoftwareConfig() when the Excel file has extra config (likely version difference)
    def testReadSoftwareConfigExtraField(self):
        ws = self.ws
        ws['A1'] = 'num_person'
        ws['B1'] = 5
        ws['A2'] = 'start_date'
        ws['B2'] =  self.test_start_date
        ws['A3'] = 'end_date'
        ws['B3'] = self.test_end_date
        ws['A4'] = 'extra field'
        ws['B4'] = 1

        self.assertRaises(TypeError, excel_input.ReadSoftwareConfig, ws)


if __name__ == '__main__':
    unittest.main()