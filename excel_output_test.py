
import data
import datetime
import openpyxl  
import unittest
import excel_output


class ExcelOutputTest(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.wb = openpyxl.Workbook()

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        cls.wb = None  # Is this line necessary?

    def setUp(self):
        self.ws = ExcelOutputTest.wb.create_sheet()

        self.test_start_date = datetime.date(2020, 1, 31)
        self.test_end_date = datetime.date(2020, 2, 1)
        self.test_config = excel_output.SoftwareConfig(
            start_date=self.test_start_date, end_date=self.test_end_date, num_person=6)

    def tearDown(self):
        ExcelOutputTest.wb.remove(self.ws)
        self.ws = None  # Is this line necessary?

    def test_empty(self):
        self.assertIsNone(None)

    # Test WriteDateConfig() with empty date constraints. Needed for barebone output
    def testWriteDateConfigEmptyConstraints(self):
        ws = self.ws
        excel_output.WriteDateConfig(ws, [], self.test_config)

        # Check header columns
        self.assertEqual(ws.cell(row=1, column=2).value, '데이 근무자 수')
        self.assertEqual(ws.cell(row=1, column=3).value, '이브닝 근무자 수')
        self.assertEqual(ws.cell(row=1, column=4).value, '나이트 근무자 수')

        # Check dates
        self.assertEqual(ws.cell(row=2, column=1).value, self.test_start_date)
        self.assertEqual(ws.cell(row=3, column=1).value, self.test_end_date)

        # Check all date config values are empty
        self.assertIsNone(ws.cell(row=2, column=2).value)
        self.assertIsNone(ws.cell(row=2, column=3).value)
        self.assertIsNone(ws.cell(row=2, column=4).value)
        self.assertIsNone(ws.cell(row=3, column=2).value)
        self.assertIsNone(ws.cell(row=3, column=3).value)
        self.assertIsNone(ws.cell(row=3, column=4).value)

     # Test WriteDateConfig()
    def testWriteDateConfig(self):
        ws = self.ws
        dc1 = data.DateConstraint(self.test_start_date, 3, 2, 1)
        dc2 = data.DateConstraint(self.test_end_date, 6, 5, 4)
        excel_output.WriteDateConfig(ws, [dc1, dc2], self.test_config)

        # Check header columns
        self.assertEqual(ws.cell(row=1, column=2).value, '데이 근무자 수')
        self.assertEqual(ws.cell(row=1, column=3).value, '이브닝 근무자 수')
        self.assertEqual(ws.cell(row=1, column=4).value, '나이트 근무자 수')

        # Check dates
        self.assertEqual(ws.cell(row=2, column=1).value, self.test_start_date)
        self.assertEqual(ws.cell(row=3, column=1).value, self.test_end_date)

        # Check all date config values are correct
        self.assertEqual(ws.cell(row=2, column=2).value, 3)
        self.assertEqual(ws.cell(row=2, column=3).value, 2)
        self.assertEqual(ws.cell(row=2, column=4).value, 1)
        self.assertEqual(ws.cell(row=3, column=2).value, 6)
        self.assertEqual(ws.cell(row=3, column=3).value, 5)
        self.assertEqual(ws.cell(row=3, column=4).value, 4)


    # Test WriteSoftwareConfig()
    def testWriteSoftwareConfig(self):
        ws = self.ws
        excel_output.WriteSoftwareConfig(ws, self.test_config)

        self.assertEqual(ws.cell(row=1, column=1).value, 'start_date')
        self.assertEqual(ws.cell(row=1, column=2).value, self.test_start_date)
        self.assertEqual(ws.cell(row=2, column=1).value, 'end_date')
        self.assertEqual(ws.cell(row=2, column=2).value, self.test_end_date)
        self.assertEqual(ws.cell(row=3, column=1).value, 'num_person')
        self.assertEqual(ws.cell(row=3, column=2).value, 6)
        # This is to ensure that only the required fields were written.
        # This should be modified when there are more config values
        self.assertIsNone(self.ws.cell(row=4, column=1).value)



if __name__ == '__main__':
    unittest.main()