import collections
import datetime

import openpyxl  

from shiftscheduler.data_types import data_types
from shiftscheduler.excel import constants
from shiftscheduler.excel import util as excel_util
from shiftscheduler.util import date_util


# Note: This module does not validate values.


# Returns assignment dict, dict of (datetime.date, str) -> data_types.ShiftType
def ReadTimetable(ws, config, start_row=1, start_col=1):
    total_dates = (config.end_date - config.start_date).days + 1
    assignment_dict = dict()

    for row_index in range(start_row+1, start_row+config.num_person+1):
        # TODO: Check that the name is not None
        name_cell = ws.cell(row=row_index, column=start_col)
        for col_index in range(start_col+1, start_col+total_dates+1):
            # TODO: Check that the date is not None
            date_cell = ws.cell(row=start_row, column=col_index) 
            shift_cell = ws.cell(row=row_index, column=col_index)
            shift_type = data_types.ShiftType.FromShortName(shift_cell.value)
            
            #if shift_type is not None:  # If the cell is not empty
            # Assign None shift type if not exists. This is needed for validation check
            assignment_dict[(excel_util.CellToDate(date_cell), name_cell.value)] = shift_type

    return assignment_dict
    

# Returns list of data_types.PersonConstraint
def ReadPersonConfig(ws, config, start_row=1, start_col=1):
    person_constraints = []
    for row_index in range(start_row+1, start_row+config.num_person+1):
        name_cell = ws.cell(row=row_index, column=start_col)
        mcw_cell = ws.cell(row=row_index, column=start_col+1)  # Max consecutive workdays
        mcn_cell = ws.cell(row=row_index, column=start_col+2)  # Max consecutive nights
        min_tw_cell = ws.cell(row=row_index, column=start_col+3)  # Min total workdays
        max_tw_cell = ws.cell(row=row_index, column=start_col+4)  # Max total workdays

        person_constraint = data_types.PersonConfig(
            name_cell.value, mcw_cell.value, mcn_cell.value, min_tw_cell.value, max_tw_cell.value)
        person_constraints.append(person_constraint)
    
    return person_constraints
    

# Returns list of data_types.DateConstraint
def ReadDateConfig(ws, config, start_row=1, start_col=1):
    date_constraints = []

    total_dates = (config.end_date - config.start_date).days + 1
    for row_index in range(start_row+1, start_row+total_dates+1):
        date_cell = ws.cell(row=row_index, column=start_col)
        day_cell = ws.cell(row=row_index, column=start_col+1)  # Number of day shift workers
        evening_cell = ws.cell(row=row_index, column=start_col+2)  # Number of evening shift workers
        night_cell = ws.cell(row=row_index, column=start_col+3)  # Number of night shift workers

        date_constraint = data_types.DateConfig(
            excel_util.CellToDate(date_cell), day_cell.value, evening_cell.value, night_cell.value)
        date_constraints.append(date_constraint)
    
    return date_constraints


# Returns Config
def ReadSoftwareConfig(ws, start_row=1, start_col=1):
    row_index = start_row
    config_dict = dict()
    while True:
        name = ws.cell(row=row_index, column=start_col).value
        if name is None:  # Reached the end of configs
            break
        value = ws.cell(row=row_index, column=start_col+1).value

        if type(value) is datetime.datetime:
            value = value.date()

        config_dict[name] = value
        row_index += 1 

    # TODO: This raises error when the Excel sheet has extra config names.
    # Should this be the default behavior?
    return data_types.SoftwareConfig(**config_dict)
    

# Read Excel file and convert to (data_types.Schedule, data_types.Assignment) objects
def ReadFromExcelFile(filepath):
    wb = openpyxl.load_workbook(filepath) 

    ws = wb[constants.SHEET_SOFTWARE_CONFIG]
    software_config = ReadSoftwareConfig(ws)
    
    ws= wb[constants.SHEET_PERSON_CONFIG]
    person_configs = ReadPersonConfig(ws, software_config)
    
    ws = wb[constants.SHEET_DATE_CONFIG] 
    date_configs = ReadDateConfig(ws, software_config)

    ws = wb[constants.SHEET_TIMETABLE]
    assignment_dict = ReadTimetable(ws, software_config)

    # Crate schedule object
    total_schedule = data_types.TotalSchedule(
        software_config=software_config,
        person_configs=person_configs,
        date_configs=date_configs,
        assignment_dict=assignment_dict,
    )

    return total_schedule

