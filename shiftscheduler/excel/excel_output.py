
import openpyxl  

from openpyxl import styles
from openpyxl.styles import fills

from openpyxl.formatting import rule as xlrule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import cell as xlcell
import date_util
import datetime
import data
import collections



# Config values for each person. The order must be preserved
CONFIG_PERSON_HEADER_NAMES = (
    '최대 근무일 (연속)',
    '최대 나이트 (연속)',
    '최소 근무일 (전체)',
    '최대 근무일 (전체)',
)


# Config values for each date. The order must be preserved
CONFIG_DATE_HEADER_NAMES = (
    '데이 근무자 수',
    '이브닝 근무자 수',
    '나이트 근무자 수',
)


# Background colors to use for timetable
COLOR_MELON = 'fdbcb4'  # Peach-like soft red
COLOR_PASTEL_YELLOW = 'fdfd96'
COLOR_AERO_BLUE = 'c9ffe5'
COLOR_LIGHT_GRAY = 'bbbbbb'


# Background fills to user for timetable
DAY_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_MELON, end_color=COLOR_MELON)
EVENING_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_AERO_BLUE, end_color=COLOR_AERO_BLUE)
NIGHT_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_PASTEL_YELLOW, end_color=COLOR_PASTEL_YELLOW)
OFF_FILL = styles.PatternFill(
    fill_type=fills.FILL_SOLID, start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY)


# If all names are unique. Prevents mistaken duplicate names
def AllNamesUnique(names):
    return len(names) == len(set(names))


# ws: Excel sheet
# schedule: data.Schedule, for start_date and end_date
# assignment: data.Assignment, for assignment
def WriteTimetable(ws, software_config, names, assignment_dict, start_row=1, start_col=1):
    start_date = software_config.start_date
    end_date = software_config.end_date
    all_dates = list(date_util.GenerateAllDates(start_date, end_date))

    end_row = start_row + len(names)  # inclusive
    end_col = start_col + len(all_dates)  # inclusive
    
    # TODO: Create custom exceptions
    if not AllNamesUnique(names):
        raise Exception('Duplicate names')
    if start_date > end_date:
        raise Exception('Start date is later than end date')
    
    # Fill the head rows with people's names
    for i, name in enumerate(names):
        c = ws.cell(row=start_row+1+i, column=start_col)
        c.value = name
    
    # Fill the head columns with dates
    for i, work_date in enumerate(all_dates):
        c = ws.cell(row=start_row, column=start_col+1+i)
        c.value = work_date
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(start_col+1+i)].width = 12  # 10 for date

    # Fill shifts from assignments
    for row_index in range(start_row+1, end_row+1):
        name = ws.cell(row=row_index, column=start_col).value
        for col_index in range(start_col+1, end_col+1):
            work_date = ws.cell(row=start_row, column=col_index).value
            cell = ws.cell(row=row_index, column=col_index)
            shift_type = assignment_dict.get((work_date, name))
            
            # If shift (including OFF) exists, write the short name.
            if shift_type is not None:
                cell.value = shift_type.ShortName()         


    # Add conditional formatting to cells
    cell_range = '%s%d:%s%d' % (
        xlcell.get_column_letter(start_col), start_row,
        xlcell.get_column_letter(end_col), end_row
    )

    # Add background colors based on value. Values are case-insensitive
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"D"'], fill=DAY_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"E"'], fill=EVENING_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"N"'], fill=NIGHT_FILL))
    ws.conditional_formatting.add(
        cell_range, xlrule.CellIsRule(operator='=', formula=['"O"'], fill=OFF_FILL))



def WritePersonConfigs(ws, person_configs, start_row=1, start_col=1):
    # Fill the head rows with people's names
    # Note that row/col indexes are 1-based in Excel    
    for i, person_config in enumerate(person_configs):
        c = ws.cell(row=start_row+1+i, column=start_col)
        c.value = person_config.name

    # Fill the head columns with config names
    for i, config_display_name in enumerate(CONFIG_PERSON_HEADER_NAMES):
        c = ws.cell(row=start_row, column=start_col+1+i)
        c.value = config_display_name
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(start_col+1+i)].width = len(config_display_name)*1.7

    # Fill config values from schedule
    name_to_config = {c.name: c for c in person_configs}
    for row_index in range(start_row+1, start_row+1+len(person_configs)):
        name = ws.cell(row=row_index, column=start_col).value
        person_config = name_to_config.get(name)
        
        # Config can be None for barebone outputs
        if person_config is not None:
            # Write config values. Iterate fields of PersonConfig namedtuple.
            for col_offset, value in enumerate(person_config._asdict().values()):
                cell = ws.cell(row=row_index, column=start_col+col_offset)
                cell.value = value   
        

def WriteDateConfigs(ws, date_configs, config, start_row=1, start_col=1):
    date_to_config = {c.work_date: c for c in date_configs}

    # Fill the head rows with dates
    # Note that row/col indexes are 1-based in Excel    
    all_dates = list(date_util.GenerateAllDates(config.start_date, config.end_date))
    for i, work_date in enumerate(all_dates):
        c = ws.cell(row=start_row+1+i, column=start_col)
        c.value = work_date

    # Fill the head columns with config names
    for i, config_display_name in enumerate(CONFIG_DATE_HEADER_NAMES):
        c = ws.cell(row=start_row, column=start_col+1+i)
        c.value = config_display_name
        # Adjust width
        ws.column_dimensions[xlcell.get_column_letter(start_col+1+i)].width = len(config_display_name)*2
    
    # Fill date configs
    for row_index in range(start_row+1, start_row+1+len(all_dates)):
        work_date = ws.cell(row=row_index, column=start_col).value
        date_config = date_to_config.get(work_date)
        
        # Config can be None for barebone outputs
        if date_config is not None:
            # Write config values. Iterate fields of DateConfig namedtuple.
            for col_offset, value in enumerate(date_config._asdict().values()):
                cell = ws.cell(row=row_index, column=start_col+col_offset)
                cell.value = value


def WriteSoftwareConfig(ws, software_config, start_row=1, start_col=1):
    # Iterate fields and values in namedtuple in order
    for row_offset, (name, value) in enumerate(software_config._asdict().items()):
        # First column is config name
        name_cell = ws.cell(row=start_row+row_offset, column=start_col)
        name_cell.value = name
        # Second column is config value
        value_cell = ws.cell(row=start_row+row_offset, column=start_col+1)
        value_cell.value = value


# Create Output Excel file
def CreateWorkbook(software_config, person_configs, date_configs, assignment_dict):
    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = '일정표'
    names = [c.name for c in person_configs]
    WriteTimetable(wb.active, software_config, names, assignment_dict)

    ws = wb.create_sheet(title='간호사별 설정')
    WritePersonConfigs(ws, person_configs)

    ws = wb.create_sheet(title='날짜별 설정')
    WriteDateConfigs(ws, date_configs, software_config)

    ws = wb.create_sheet(title='Config')
    WriteSoftwareConfig(ws, software_config)

    return wb

def FromTotalSchedule(total_schedule, filepath):
    wb = CreateWorkbook(
        total_schedule.software_config,
        total_schedule.person_configs,
        total_schedule.date_configs,
        total_schedule.assignment_dict)
    wb.save(filepath)