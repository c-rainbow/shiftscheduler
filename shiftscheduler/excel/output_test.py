
import datetime
import unittest

import openpyxl  

from shiftscheduler.data_types import data_types
from shiftscheduler.excel import output as excel_output



class ExcelOutputTest(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.wb = openpyxl.Workbook()

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        cls.wb = None  # Is this line necessary?

    def setUp(self):
        self.ws = ExcelOutputTest.wb.create_sheet()
        self.test_start_date = datetime.date(2020, 1, 31)
        self.test_end_date = datetime.date(2020, 2, 1)
        self.test_config = data_types.SoftwareConfig(
            start_date=self.test_start_date, end_date=self.test_end_date, num_person=6)

    def tearDown(self):
        ExcelOutputTest.wb.remove(self.ws)
        self.ws = None  # Is this line necessary?

    # Test WriteTimetable() with empty shift values. Needed for barebone output
    def testWriteTimetableEmpty(self):
        ws = self.ws
        excel_output.WriteTimetable(ws, self.test_config, ('간호사1', '간호사2'), dict())
        
        # Check header columns
        self.assertEqual(ws.cell(row=1, column=2).value, self.test_start_date)
        self.assertEqual(ws.cell(row=1, column=3).value, self.test_end_date)

        # Check header rows
        self.assertEqual(ws.cell(row=2, column=1).value, '간호사1')
        self.assertEqual(ws.cell(row=3, column=1).value, '간호사2')
        
        # Check the timetable is empty
        self.assertIsNone(ws.cell(row=2, column=2).value)
        self.assertIsNone(ws.cell(row=2, column=3).value)
        self.assertIsNone(ws.cell(row=3, column=2).value)
        self.assertIsNone(ws.cell(row=3, column=3).value)

    # Test WriteTimetable()
    def testWriteTimetable(self):
        ws = self.ws
        names = ('간호사1', '간호사2', '간호사3')
        assignment_dict = {
            (self.test_start_date, '간호사1'): data_types.ShiftType.DAY,
            (self.test_end_date, '간호사1'): data_types.ShiftType.DAY,
            (self.test_start_date, '간호사2'): data_types.ShiftType.EVENING,
            (self.test_end_date, '간호사2'): data_types.ShiftType.OFF,
            # Shift of 간호사3 is undecided on the start date.
            (self.test_end_date, '간호사3'): data_types.ShiftType.NIGHT,
        }
        excel_output.WriteTimetable(ws, self.test_config, names, assignment_dict)
        
        # Check header columns
        self.assertEqual(ws.cell(row=1, column=2).value, self.test_start_date)
        self.assertEqual(ws.cell(row=1, column=3).value, self.test_end_date)
        self.assertIsNone(ws.cell(row=1, column=4).value)  # Header columns should end here

        # Check header rows
        self.assertEqual(ws.cell(row=2, column=1).value, '간호사1')
        self.assertEqual(ws.cell(row=3, column=1).value, '간호사2')
        self.assertEqual(ws.cell(row=4, column=1).value, '간호사3')
        self.assertIsNone(ws.cell(row=5, column=1).value)  # Header rows should end here
        
        # Check timetable of person 1
        self.assertEqual(ws.cell(row=2, column=2).value, 'D')
        self.assertEqual(ws.cell(row=2, column=3).value, 'D')
        self.assertIsNone(ws.cell(row=2, column=4).value)

        # Check timetable of person 2
        self.assertEqual(ws.cell(row=3, column=2).value, 'E')
        self.assertEqual(ws.cell(row=3, column=3).value, 'O')
        self.assertIsNone(ws.cell(row=3, column=4).value)

        # Check timetable of person 3
        self.assertIsNone(ws.cell(row=4, column=2).value)
        self.assertEqual(ws.cell(row=4, column=3).value, 'N')
        self.assertIsNone(ws.cell(row=4, column=4).value)

        # Nothing should be written after person 3
        self.assertIsNone(ws.cell(row=5, column=2).value)

    # Test WritePersonConfig() with empty constraints. Needed for barebone output
    def testWritePersonConfigEmptyConstraints(self):
        ws = self.ws
        pc1 = data_types.PersonConfig('간호사1', None, None, None, None)
        excel_output.WritePersonConfigs(ws, [pc1])

        # Check header columns
        self.assertEqual(ws.cell(row=1, column=2).value, '최대 근무일 (연속)')
        self.assertEqual(ws.cell(row=1, column=3).value, '최대 나이트 (연속)')
        self.assertEqual(ws.cell(row=1, column=4).value, '최소 근무일 (전체)')
        self.assertEqual(ws.cell(row=1, column=5).value, '최대 근무일 (전체)')
        self.assertIsNone(ws.cell(row=1, column=6).value)  # Header columns should end here

        # Check header rows
        self.assertEqual(ws.cell(row=2, column=1).value, '간호사1')
        self.assertIsNone(ws.cell(row=3, column=1).value)  # Header rows should end here

        # Check all person constraint values are empty
        self.assertIsNone(ws.cell(row=2, column=2).value)
        self.assertIsNone(ws.cell(row=2, column=3).value)
        self.assertIsNone(ws.cell(row=2, column=4).value)
        self.assertIsNone(ws.cell(row=2, column=5).value)

    # Test WritePersonConfig() with empty constraints. Needed for barebone output
    def testWritePersonConfig(self):
        ws = self.ws
        pc1 = data_types.PersonConfig('간호사1', 5, 3, 25, 28)
        pc2 = data_types.PersonConfig('간호사2', 6, 4, 22, 26)
        excel_output.WritePersonConfigs(ws, [pc1, pc2])

        # Check header columns
        self.assertEqual(ws.cell(row=1, column=2).value, '최대 근무일 (연속)')
        self.assertEqual(ws.cell(row=1, column=3).value, '최대 나이트 (연속)')
        self.assertEqual(ws.cell(row=1, column=4).value, '최소 근무일 (전체)')
        self.assertEqual(ws.cell(row=1, column=5).value, '최대 근무일 (전체)')
        self.assertIsNone(ws.cell(row=1, column=6).value)  # Header columns should end here

        # Check header rows
        self.assertEqual(ws.cell(row=2, column=1).value, '간호사1')
        self.assertEqual(ws.cell(row=3, column=1).value, '간호사2')
        self.assertIsNone(ws.cell(row=4, column=1).value)  # Header rows should end here

        # Check all person constraint values are correct
        self.assertEqual(ws.cell(row=2, column=2).value, 5)
        self.assertEqual(ws.cell(row=2, column=3).value, 3)
        self.assertEqual(ws.cell(row=2, column=4).value, 25)
        self.assertEqual(ws.cell(row=2, column=5).value, 28)
        self.assertIsNone(ws.cell(row=2, column=6).value)

        self.assertEqual(ws.cell(row=3, column=2).value, 6)
        self.assertEqual(ws.cell(row=3, column=3).value, 4)
        self.assertEqual(ws.cell(row=3, column=4).value, 22)
        self.assertEqual(ws.cell(row=3, column=5).value, 26)
        self.assertIsNone(ws.cell(row=3, column=6).value)

        self.assertIsNone(ws.cell(row=4, column=2).value)
        
    # Test WriteDateConfig() with empty date constraints. Needed for barebone output
    def testWriteDateConfigEmptyConstraints(self):
        ws = self.ws
        excel_output.WriteDateConfigs(ws, [], self.test_config)

        # Check header columns
        self.assertEqual(ws.cell(row=1, column=2).value, '데이 근무자 수')
        self.assertEqual(ws.cell(row=1, column=3).value, '이브닝 근무자 수')
        self.assertEqual(ws.cell(row=1, column=4).value, '나이트 근무자 수')

        # Check dates
        self.assertEqual(ws.cell(row=2, column=1).value, self.test_start_date)
        self.assertEqual(ws.cell(row=3, column=1).value, self.test_end_date)

        # Check all date config values are empty
        self.assertIsNone(ws.cell(row=2, column=2).value)
        self.assertIsNone(ws.cell(row=2, column=3).value)
        self.assertIsNone(ws.cell(row=2, column=4).value)
        self.assertIsNone(ws.cell(row=3, column=2).value)
        self.assertIsNone(ws.cell(row=3, column=3).value)
        self.assertIsNone(ws.cell(row=3, column=4).value)

     # Test WriteDateConfig()
    def testWriteDateConfig(self):
        ws = self.ws
        dc1 = data_types.DateConfig(self.test_start_date, 3, 2, 1)
        dc2 = data_types.DateConfig(self.test_end_date, 6, 5, 4)
        excel_output.WriteDateConfigs(ws, [dc1, dc2], self.test_config)

        # Check header columns
        self.assertEqual(ws.cell(row=1, column=2).value, '데이 근무자 수')
        self.assertEqual(ws.cell(row=1, column=3).value, '이브닝 근무자 수')
        self.assertEqual(ws.cell(row=1, column=4).value, '나이트 근무자 수')

        # Check dates
        self.assertEqual(ws.cell(row=2, column=1).value, self.test_start_date)
        self.assertEqual(ws.cell(row=3, column=1).value, self.test_end_date)

        # Check all date config values are correct
        self.assertEqual(ws.cell(row=2, column=2).value, 3)
        self.assertEqual(ws.cell(row=2, column=3).value, 2)
        self.assertEqual(ws.cell(row=2, column=4).value, 1)
        self.assertEqual(ws.cell(row=3, column=2).value, 6)
        self.assertEqual(ws.cell(row=3, column=3).value, 5)
        self.assertEqual(ws.cell(row=3, column=4).value, 4)

    # Test WriteSoftwareConfig()
    def testWriteSoftwareConfig(self):
        ws = self.ws
        excel_output.WriteSoftwareConfig(ws, self.test_config)

        self.assertEqual(ws.cell(row=1, column=1).value, 'start_date')
        self.assertEqual(ws.cell(row=1, column=2).value, self.test_start_date)
        self.assertEqual(ws.cell(row=2, column=1).value, 'end_date')
        self.assertEqual(ws.cell(row=2, column=2).value, self.test_end_date)
        self.assertEqual(ws.cell(row=3, column=1).value, 'num_person')
        self.assertEqual(ws.cell(row=3, column=2).value, 6)
        # This is to ensure that only the required fields were written.
        # This should be modified when there are more config values
        self.assertIsNone(self.ws.cell(row=4, column=1).value)


if __name__ == '__main__':
    unittest.main()