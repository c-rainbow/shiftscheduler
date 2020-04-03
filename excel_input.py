import collections
import openpyxl  

from openpyxl import styles
from openpyxl.styles import fills

from openpyxl.formatting import rule as xlrule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import cell as xlcell
import date_util
import datetime
import data
# from nurse_scheduling.excel import constants


# Note: This module does not validate values.


# Returns assignment dict, dict of (datetime.date, str) -> data.ShiftType
def ReadTimetable(ws, config, start_row=1, start_col=1):
    total_dates = (config.end_date - config.start_date).days + 1
    assignment_dict = dict()

    for row_index in range(start_row+1, start_row+config.num_person+1):
        # TODO: Check that the name is not None
        name_cell = ws.cell(row=row_index, column=start_col)
        for col_index in range(start_col+1, start_col+total_dates+1):
            # TODO: Check that the date is not None
            date_cell = ws.cell(row=start_row, column=col_index) 
            shift_cell = ws.cell(row=row_index, column=col_index)
            shift_type = data.ShiftType.FromShortName(shift_cell.value)
            
            if shift_type is not None:  # If the cell is not empty
                assignment_dict[(datetime.date.fromisoformat(date_cell.value), name_cell.value)] = shift_type

    return assignment_dict
    

# Returns list of data.PersonConstraint
def ReadPersonConfig(ws, config, start_row=1, start_col=1):
    person_constraints = []
    for row_index in range(start_row+1, start_row+config.num_person+1):
        name_cell = ws.cell(row=row_index, column=start_col)
        mcw_cell = ws.cell(row=row_index, column=start_col+1)  # Max consecutive workdays
        mcn_cell = ws.cell(row=row_index, column=start_col+2)  # Max consecutive nights
        min_tw_cell = ws.cell(row=row_index, column=start_col+3)  # Min total workdays
        max_tw_cell = ws.cell(row=row_index, column=start_col+4)  # Max total workdays

        person_constraint = data.PersonConfig(
            name_cell.value, mcw_cell.value, mcn_cell.value, min_tw_cell.value, max_tw_cell.value)
        person_constraints.append(person_constraint)
    
    return person_constraints
    

# Returns list of data.DateConstraint
def ReadDateConfig(ws, config, start_row=1, start_col=1):
    date_constraints = []

    total_dates = (config.end_date - config.start_date).days + 1
    for row_index in range(start_row+1, start_row+total_dates+1):
        date_cell = ws.cell(row=row_index, column=start_col)
        day_cell = ws.cell(row=row_index, column=start_col+1)  # Number of day shift workers
        evening_cell = ws.cell(row=row_index, column=start_col+2)  # Number of evening shift workers
        night_cell = ws.cell(row=row_index, column=start_col+3)  # Number of night shift workers

        date_constraint = data.DateConfig(
            datetime.date.fromisoformat(date_cell.value), day_cell.value, evening_cell.value, night_cell.value)
        date_constraints.append(date_constraint)
    
    return date_constraints


# Returns Config
def ReadSoftwareConfig(ws, start_row=1, start_col=1):
    row_index = start_row
    config_dict = dict()
    while True:
        name = ws.cell(row=row_index, column=start_col).value
        if name is None:  # Reached the end of configs
            break
        value = ws.cell(row=row_index, column=start_col+1).value

        if type(value) is datetime.datetime:
            value = value.date()

        config_dict[name] = value
        row_index += 1 

    # TODO: This raises error when the Excel sheet has extra config names.
    # Should this be the default behavior?
    return data.SoftwareConfig(**config_dict)
    

# Read Excel file and convert to (data.Schedule, data.Assignment) objects
def ReadFromExcelFile(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True) 

    ws = wb['Config']
    software_config = ReadSoftwareConfig(ws)
    
    ws= wb['간호사별 설정']
    person_configs = ReadPersonConfig(ws, software_config)
    
    ws = wb['날짜별 설정'] 
    date_configs = ReadDateConfig(ws, software_config)

    ws = wb['일정표']
    assignment_dict = ReadTimetable(ws, software_config)

    # Crate schedule object
    total_schedule = data.TotalSchedule(
        software_config=software_config,
        person_configs=person_configs,
        date_configs=date_configs,
        assignment_dict=assignment_dict,
    )

    return total_schedule



if __name__ == '__main__':
    schedule, assignments = ReadFromExcelFile('tmp/sample.xlsx')
    print(schedule)
    print(assignments)